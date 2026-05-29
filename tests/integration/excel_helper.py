"""
Excel 集成测试辅助模块

负责:
- 创建/清理测试用 Excel 文件
- 启动/关闭 Excel 进程
- 读取 Excel 单元格内容用于断言验证
- 等待窗口就绪
"""

import os
import sys
import time
import ctypes
import tempfile
import subprocess
from typing import Optional

# 将项目根目录加入 sys.path
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import win32gui
    import win32con
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

try:
    from rpa.execution import find_window_by_title
except ImportError:
    def find_window_by_title(title: str) -> Optional[int]:
        """回退实现"""
        if not HAS_WIN32:
            return None
        def _enum(hwnd, _):
            try:
                if win32gui.IsWindowVisible(hwnd):
                    t = win32gui.GetWindowText(hwnd)
                    if title.lower() in t.lower():
                        _enum.found = hwnd
                        return False
            except Exception:
                pass
            return True
        _enum.found = None
        win32gui.EnumWindows(_enum, None)
        return _enum.found


# ============================================================
# 测试 Excel 文件生成
# ============================================================

def create_test_excel(filepath: str, sheet_name: str = "Sheet",
                      data: list = None) -> str:
    """
    创建一个测试用 .xlsx 文件。

    Args:
        filepath: 目标文件路径
        sheet_name: 工作表名
        data: 二维数据列表，如 [["A1","B1"],["A2","B2"]]

    Returns:
        文件路径
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("需要 openpyxl 库，请执行: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if data:
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(filepath)
    return filepath


def read_excel_cell(filepath: str, cell: str, sheet_name: str = "Sheet") -> str:
    """
    读取 Excel 单元格值。

    Args:
        filepath: Excel 文件路径
        cell: 单元格坐标如 "A1"
        sheet_name: 工作表名

    Returns:
        单元格文本值（None 返回空字符串）
    """
    if not HAS_OPENPYXL:
        raise RuntimeError("需要 openpyxl 库")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    value = ws[cell].value
    wb.close()
    return str(value) if value is not None else ""


def read_excel_row(filepath: str, row: int, max_col: int = 10,
                   sheet_name: str = "Sheet") -> list:
    """读取 Excel 一整行数据。"""
    if not HAS_OPENPYXL:
        raise RuntimeError("需要 openpyxl 库")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    values = []
    for col in range(1, max_col + 1):
        v = ws.cell(row=row, column=col).value
        values.append(str(v) if v is not None else "")
    wb.close()
    return values


# ============================================================
# Excel 进程管理
# ============================================================

EXCEL_PATHS = [
    r"C:\Program Files\Microsoft Office\root\Office16\EXCEL.EXE",
    r"C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE",
    r"C:\Program Files\Microsoft Office\Office16\EXCEL.EXE",
    r"C:\Program Files (x86)\Microsoft Office\Office16\EXCEL.EXE",
]

# WPS Office 路径（国产化兼容）
WPS_PATHS = [
    r"C:\Users\{}\AppData\Local\Kingsoft\WPS Office\*\office6\et.exe",
    r"C:\Program Files (x86)\Kingsoft\WPS Office\*\office6\et.exe",
    r"C:\Program Files\Kingsoft\WPS Office\*\office6\et.exe",
]

# 存储找到的可执行文件路径（缓存）
_excel_exe_cache: Optional[str] = None


def _get_excel_window_suffix() -> str:
    """
    根据可执行文件名推断窗口标题后缀。
    Excel: " - Excel" 或 " - Microsoft Excel"
    WPS:   " - WPS Office" 或 " - WPS表格"
    """
    exe = find_excel_exe()
    name = os.path.basename(exe).lower()
    if name.startswith('et'):
        return "WPS"
    return "Excel"


def get_window_title_fragment(filepath: str) -> str:
    """
    根据文件路径返回用于窗口查找的标题片段。
    不同办公软件在标题栏显示的文件名格式不同。

    Args:
        filepath: 文件路径

    Returns:
        窗口标题中可唯一识别该文件的部分
    """
    return os.path.basename(filepath)

EXCEL_REGISTRY_KEYS = [
    # Microsoft Excel (Office 2016/2019/365 Click-to-Run)
    (r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\excel.exe", ""),
    # Microsoft Office 标准安装
    (r"SOFTWARE\Microsoft\Office\16.0\Excel\InstallRoot", "Path"),
    (r"SOFTWARE\Microsoft\Office\15.0\Excel\InstallRoot", "Path"),
    (r"SOFTWARE\WOW6432Node\Microsoft\Office\16.0\Excel\InstallRoot", "Path"),
    (r"SOFTWARE\WOW6432Node\Microsoft\Office\15.0\Excel\InstallRoot", "Path"),
]


def _find_via_registry() -> Optional[str]:
    """通过 Windows 注册表查找 Excel 路径。"""
    try:
        import winreg
    except ImportError:
        return None

    for key_path, value_name in EXCEL_REGISTRY_KEYS:
        try:
            for hive in (winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER):
                try:
                    key = winreg.OpenKey(hive, key_path)
                    val, _ = winreg.QueryValueEx(key, value_name or "")
                    winreg.CloseKey(key)
                    if value_name:
                        exe = os.path.join(val, "EXCEL.EXE")
                    else:
                        exe = val
                    if os.path.isfile(exe):
                        return exe
                except (OSError, FileNotFoundError):
                    continue
        except Exception:
            continue
    return None


def _find_wps() -> Optional[str]:
    """查找 WPS Office 可执行文件。"""
    import glob
    username = os.environ.get("USERNAME", os.environ.get("USER", ""))
    for pattern in WPS_PATHS:
        expanded = pattern.format(username) if "{}" in pattern else pattern
        matches = glob.glob(expanded)
        for m in matches:
            if os.path.isfile(m):
                return m
    return None


def find_excel_exe() -> str:
    """
    查找 Excel/WPS 可执行文件路径。

    查找顺序:
    1. Windows 注册表
    2. 硬编码路径
    3. WPS Office

    Raises:
        FileNotFoundError: 未找到
    """
    # 1. 注册表查找
    exe = _find_via_registry()
    if exe:
        return exe

    # 2. 硬编码路径
    for p in EXCEL_PATHS:
        if os.path.exists(p):
            return p

    # 3. WPS Office
    exe = _find_wps()
    if exe:
        return exe

    raise FileNotFoundError(
        "未找到 Excel/WPS 可执行文件。\n"
        "搜索方式: 注册表 + 硬编码路径 + WPS Office\n"
        "硬编码搜索路径: " + ", ".join(EXCEL_PATHS)
    )


def is_excel_available() -> bool:
    """检查系统是否安装了 Excel/WPS。"""
    try:
        find_excel_exe()
        return True
    except FileNotFoundError:
        return False


def launch_excel(filepath: str, visible: bool = True,
                 timeout: float = 15.0) -> subprocess.Popen:
    """
    启动 Excel/WPS 并打开指定文件。

    Args:
        filepath: 要打开的文件路径 (.xlsx)
        visible: 是否可见
        timeout: 等待就绪的超时时间

    Returns:
        subprocess.Popen 对象
    """
    excel_exe = find_excel_exe()
    abs_path = os.path.abspath(filepath)

    proc = subprocess.Popen(
        [excel_exe, abs_path],
        shell=False,
    )

    # 等待窗口出现（用文件名片段匹配）
    title_fragment = get_window_title_fragment(filepath)
    _wait_for_window(title_fragment, timeout)
    return proc


def kill_excel(proc: Optional[subprocess.Popen] = None) -> None:
    """关闭 Excel 进程（强制终止）。"""
    if proc is not None:
        try:
            proc.terminate()
            proc.wait(timeout=5)
        except Exception:
            try:
                proc.kill()
            except Exception:
                pass


def _wait_for_window(title_part: str, timeout: float = 15.0) -> Optional[int]:
    """
    等待标题包含指定字符串的窗口出现。

    Args:
        title_part: 窗口标题的部分字符串
        timeout: 超时秒数

    Returns:
        窗口句柄，超时返回 None
    """
    deadline = time.time() + timeout
    while time.time() < deadline:
        hwnd = find_window_by_title(title_part)
        if hwnd is not None:
            return hwnd
        time.sleep(0.5)
    return None


def get_excel_hwnd(title_part: str, timeout: float = 3.0) -> Optional[int]:
    """
    获取 Excel 窗口句柄。

    Args:
        title_part: 窗口标题部分字符串
        timeout: 超时秒数
    """
    return _wait_for_window(title_part, timeout)


def activate_excel_window(title_part: str) -> bool:
    """
    将 Excel 窗口激活到前台。

    Args:
        title_part: 窗口标题部分字符串

    Returns:
        是否成功
    """
    hwnd = find_window_by_title(title_part)
    if hwnd is None:
        return False
    try:
        import win32gui
        win32gui.SetForegroundWindow(hwnd)
        time.sleep(0.3)
        return True
    except Exception:
        return False


def wait_ready(title_part: str, timeout: float = 3.0) -> bool:
    """等待 Excel 窗口出现并激活。"""
    hwnd = _wait_for_window(title_part, timeout)
    return hwnd is not None
